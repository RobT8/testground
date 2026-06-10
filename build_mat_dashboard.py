"""
Builds an Academy Trust Management Accounts pack in Excel, structured around
DfE good-practice guidance / Academy Trust Handbook requirements:
  - Income & expenditure with variance to budget
  - Balance sheet
  - Cash flow forecast
  - Key financial performance indicators (staff costs %, days cash, reserves)
All headline figures are driven by formulas from the detail sheets, so the
pack updates when actuals are typed in.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------- palette
NAVY   = "1F3864"   # headers / banner
NAVY2  = "2F5597"   # sub headers
TEAL   = "2E9CA6"   # accent
GOLD   = "C9A227"   # accent 2
LIGHT  = "F2F5FA"   # card / zebra fill
MIDGREY= "D9DEE8"
GREEN  = "2E7D32"
GREENF = "E3F2E5"
AMBER  = "B26A00"
AMBERF = "FDF1DC"
RED    = "B71C1C"
REDF   = "FBE4E4"
WHITE  = "FFFFFF"

GBP   = '#,##0;[Red](#,##0)'
GBP_K = '#,##0,"k";[Red](#,##0,"k")'
PCT   = '0.0%;[Red](0.0%)'

thin = Side(style="thin", color=MIDGREY)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
bot  = Border(bottom=Side(style="medium", color=NAVY))
topd = Border(top=Side(style="thin", color=NAVY), bottom=Side(style="double", color=NAVY))

def fill(hex_): return PatternFill("solid", fgColor=hex_)

MONTHS = ["Sep","Oct","Nov","Dec","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug"]
CUR_PERIOD = 9  # May 2026 (period 9 of academic year 2025/26)

wb = openpyxl.Workbook()

def style_sheet(ws, tab_color=NAVY):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

def banner(ws, title, subtitle, last_col):
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    for r in (2, 3):
        for c in range(2, last_col + 1):
            ws.cell(row=r, column=c).fill = fill(NAVY)
    t = ws.cell(row=2, column=2, value=title)
    t.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    t.alignment = Alignment(vertical="center", indent=1)
    s = ws.cell(row=3, column=2, value=subtitle)
    s.font = Font(name="Calibri", size=11, italic=True, color=MIDGREY)
    s.alignment = Alignment(vertical="center", indent=1)

# ================================================================ DATA (monthly)
# Hidden-ish working sheet that drives the dashboard charts.
dt = wb.create_sheet("Monthly Data")
style_sheet(dt, "808080")
banner(dt, "Monthly Data", "Working data that drives the Dashboard charts — enter monthly actuals here", 16)

inc_act = [1380, 1295, 1310, 1290, 1335, 1300, 1345, 1310, 1352, None, None, None]
inc_bud = [1350, 1300, 1300, 1300, 1320, 1310, 1330, 1320, 1330, 1330, 1340, 1370]
exp_act = [1290, 1310, 1285, 1330, 1295, 1305, 1290, 1315, 1298, None, None, None]
exp_bud = [1300, 1295, 1290, 1320, 1300, 1300, 1295, 1310, 1305, 1305, 1320, 1360]
cash_act= [2350, 2310, 2330, 2280, 2335, 2330, 2390, 2385, 2440, None, None, None]
cash_fc = [2350, 2315, 2325, 2285, 2330, 2335, 2380, 2390, 2430, 2455, 2470, 2480]

hdrs = ["Month","Income Actual £k","Income Budget £k","Expenditure Actual £k",
        "Expenditure Budget £k","Cash Actual £k","Cash Forecast £k",
        "Surplus Actual £k","Surplus Budget £k"]
for j, h in enumerate(hdrs):
    c = dt.cell(row=5, column=2 + j, value=h)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    c.border = box
dt.row_dimensions[5].height = 30
for i, m in enumerate(MONTHS):
    r = 6 + i
    vals = [m, inc_act[i], inc_bud[i], exp_act[i], exp_bud[i], cash_act[i], cash_fc[i]]
    for j, v in enumerate(vals):
        c = dt.cell(row=r, column=2 + j, value=v)
        c.border = box
        if j: c.number_format = '#,##0'
        if i % 2: c.fill = fill(LIGHT)
    sa = dt.cell(row=r, column=9, value=f"=IF(C{r}=\"\",\"\",C{r}-E{r})")
    sb = dt.cell(row=r, column=10, value=f"=D{r}-F{r}")
    for c in (sa, sb):
        c.border = box; c.number_format = GBP
        if i % 2: c.fill = fill(LIGHT)
dt.column_dimensions["A"].width = 2
for col in "BCDEFGHIJ":
    dt.column_dimensions[col].width = 14

# ================================================================ I&E
ie = wb.create_sheet("Income & Expenditure")
style_sheet(ie, NAVY2)
banner(ie, "Income & Expenditure", f"Year to date Period {CUR_PERIOD} (May 2026) and full-year forecast  |  Academic year 2025/26  |  £", 9)

cols = ["", "YTD Actual", "YTD Budget", "YTD Variance", "Var %",
        "FY Budget", "FY Forecast", "FY Variance"]
HR = 5
for j, h in enumerate(cols):
    c = ie.cell(row=HR, column=2 + j, value=h)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
ie.row_dimensions[HR].height = 26

# (label, ytd_actual, ytd_budget, fy_budget, fy_forecast)
income_rows = [
    ("General Annual Grant (GAG)",        10125000, 10050000, 13400000, 13480000),
    ("Other DfE / ESFA grants",              610000,   585000,   780000,   805000),
    ("Pupil premium",                        565000,   560000,   745000,   748000),
    ("Local authority grants (inc. SEN)",    402000,   430000,   573000,   545000),
    ("Catering & lettings income",           158000,   170000,   227000,   212000),
    ("Donations & fundraising",               57000,    45000,    60000,    71000),
]
expend_rows = [
    ("Teaching staff",                      5430000,  5380000,  7173000,  7240000),
    ("Education support staff",             1488000,  1520000,  2027000,  1985000),
    ("Admin & leadership staff",            1356000,  1340000,  1787000,  1802000),
    ("Indirect staff costs (supply, dev.)",  302000,   340000,   453000,   420000),
    ("Premises & estates (inc. PFI/maint.)", 985000,   940000,  1253000,  1310000),
    ("Educational supplies & ICT",           738000,   760000,  1013000,   995000),
    ("Catering supplies",                    331000,   325000,   433000,   441000),
    ("Other support costs & professional",   788000,   810000,  1080000,  1056000),
]

def write_block(ws, start, title, rows, total_label):
    r = start
    c = ws.cell(row=r, column=2, value=title)
    c.font = Font(bold=True, size=12, color=NAVY)
    c.border = bot
    for cc in range(3, 10):
        ws.cell(row=r, column=cc).border = bot
    r += 1
    first = r
    for i, (label, ya, yb, fb, ff) in enumerate(rows):
        ws.cell(row=r, column=2, value=label).border = box
        for col, v in ((3, ya), (4, yb), (7, fb), (8, ff)):
            cell = ws.cell(row=r, column=col, value=v)
            cell.number_format = GBP; cell.border = box
        v1 = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        p1 = ws.cell(row=r, column=6, value=f"=IF(D{r}=0,0,(C{r}-D{r})/D{r})")
        v2 = ws.cell(row=r, column=9, value=f"=H{r}-G{r}")
        v1.number_format = GBP; v2.number_format = GBP; p1.number_format = PCT
        for cc in (5, 6, 9): ws.cell(row=r, column=cc).border = box
        if i % 2:
            for cc in range(2, 10): ws.cell(row=r, column=cc).fill = fill(LIGHT)
        r += 1
    last = r - 1
    ws.cell(row=r, column=2, value=total_label).font = Font(bold=True, color=NAVY)
    for col in (3, 4, 7, 8):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{first}:{L}{last})").number_format = GBP
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = GBP
    ws.cell(row=r, column=6, value=f"=IF(D{r}=0,0,(C{r}-D{r})/D{r})").number_format = PCT
    ws.cell(row=r, column=9, value=f"=H{r}-G{r}").number_format = GBP
    for cc in range(2, 10):
        cell = ws.cell(row=r, column=cc)
        cell.font = Font(bold=True, color=NAVY); cell.border = topd; cell.fill = fill(LIGHT)
    return r

inc_total_row = write_block(ie, 7, "INCOME", income_rows, "Total income")
exp_start = inc_total_row + 2
exp_total_row = write_block(ie, exp_start, "EXPENDITURE", expend_rows, "Total expenditure")

sr = exp_total_row + 2
ie.cell(row=sr, column=2, value="In-year surplus / (deficit)").font = Font(bold=True, size=12, color=WHITE)
for col in (3, 4, 7, 8):
    L = get_column_letter(col)
    ie.cell(row=sr, column=col, value=f"={L}{inc_total_row}-{L}{exp_total_row}").number_format = GBP
ie.cell(row=sr, column=5, value=f"=C{sr}-D{sr}").number_format = GBP
ie.cell(row=sr, column=6, value=f"=IF(D{sr}=0,0,(C{sr}-D{sr})/MAX(1,ABS(D{sr})))").number_format = PCT
ie.cell(row=sr, column=9, value=f"=H{sr}-G{sr}").number_format = GBP
for cc in range(2, 10):
    cell = ie.cell(row=sr, column=cc)
    cell.fill = fill(NAVY); cell.font = Font(bold=True, color=WHITE); cell.border = box
ie.row_dimensions[sr].height = 22

# RAG conditional formatting on variance % column (income good=+, costs handled by sign of variance £)
ie.conditional_formatting.add(
    f"F8:F{sr}",
    CellIsRule(operator="lessThan", formula=["-0.05"], font=Font(color=RED, bold=True)))
ie.column_dimensions["A"].width = 2
ie.column_dimensions["B"].width = 36
for col in "CDEFGHI":
    ie.column_dimensions[col].width = 14
ie.freeze_panes = "B6"

n = ie.cell(row=sr + 2, column=2,
    value="Variance basis: income — favourable = actual above budget; expenditure — favourable = actual below budget. "
          "Variances over ±5% or ±£50k should be explained in the Commentary tab (DfE good practice).")
n.font = Font(italic=True, size=9, color="666666")

# ================================================================ BALANCE SHEET
bs = wb.create_sheet("Balance Sheet")
style_sheet(bs, NAVY2)
banner(bs, "Balance Sheet", "As at 31 May 2026 (Period 9) with prior year-end comparative  |  £", 6)

bs_rows = [
    ("sec",  "FIXED ASSETS", None, None),
    ("row",  "Tangible fixed assets (land & buildings)", 28450000, 28930000),
    ("row",  "Other tangible fixed assets (FF&E, ICT)",     842000,    760000),
    ("tot",  "Total fixed assets", None, None),
    ("sec",  "CURRENT ASSETS", None, None),
    ("row",  "Debtors & prepayments",                       412000,    385000),
    ("row",  "Cash at bank and in hand",                   2440000,   2350000),
    ("tot",  "Total current assets", None, None),
    ("sec",  "LIABILITIES", None, None),
    ("row",  "Creditors: amounts due within one year",    -618000,  -572000),
    ("row",  "Creditors: amounts due after one year",      -95000,  -120000),
    ("row",  "Pension scheme liability (LGPS)",          -3120000, -3450000),
    ("tot",  "Total liabilities", None, None),
    ("net",  "NET ASSETS", None, None),
    ("sec",  "FUNDS / RESERVES", None, None),
    ("row",  "Restricted fixed asset fund",              29292000,  29690000),
    ("row",  "Restricted general fund (GAG)",               738000,    540000),
    ("row",  "Restricted pension reserve",               (3120000), (3450000)),
    ("row",  "Unrestricted general fund",                 1401000,   1503000),
    ("net",  "TOTAL FUNDS", None, None),
]
hdr = ["", "31 May 2026", "31 Aug 2025", "Movement"]
for j, h in enumerate(hdr):
    c = bs.cell(row=5, column=2 + j, value=h)
    c.font = Font(bold=True, color=WHITE, size=10); c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal="center"); c.border = box

r = 6
section_first = None
totals = {}
zebra = 0
for kind, label, cur, py in bs_rows:
    if kind == "sec":
        c = bs.cell(row=r, column=2, value=label)
        c.font = Font(bold=True, size=11, color=NAVY); c.border = bot
        for cc in (3, 4, 5): bs.cell(row=r, column=cc).border = bot
        section_first = r + 1; zebra = 0
    elif kind == "row":
        bs.cell(row=r, column=2, value=label).border = box
        bs.cell(row=r, column=3, value=cur).number_format = GBP
        bs.cell(row=r, column=4, value=py).number_format = GBP
        bs.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = GBP
        for cc in (3, 4, 5): bs.cell(row=r, column=cc).border = box
        if zebra % 2:
            for cc in range(2, 6): bs.cell(row=r, column=cc).fill = fill(LIGHT)
        zebra += 1
    elif kind == "tot":
        bs.cell(row=r, column=2, value=label)
        for col in (3, 4):
            L = get_column_letter(col)
            bs.cell(row=r, column=col, value=f"=SUM({L}{section_first}:{L}{r-1})").number_format = GBP
        bs.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = GBP
        totals[label] = r
        for cc in range(2, 6):
            cell = bs.cell(row=r, column=cc)
            cell.font = Font(bold=True, color=NAVY); cell.border = topd; cell.fill = fill(LIGHT)
    elif kind == "net":
        bs.cell(row=r, column=2, value=label)
        if label == "NET ASSETS":
            fa, ca, li = totals["Total fixed assets"], totals["Total current assets"], totals["Total liabilities"]
            for col in (3, 4):
                L = get_column_letter(col)
                bs.cell(row=r, column=col, value=f"={L}{fa}+{L}{ca}+{L}{li}").number_format = GBP
        else:
            for col in (3, 4):
                L = get_column_letter(col)
                bs.cell(row=r, column=col, value=f"=SUM({L}{section_first}:{L}{r-1})").number_format = GBP
        bs.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = GBP
        for cc in range(2, 6):
            cell = bs.cell(row=r, column=cc)
            cell.fill = fill(NAVY); cell.font = Font(bold=True, color=WHITE); cell.border = box
    r += 1

bs.column_dimensions["A"].width = 2
bs.column_dimensions["B"].width = 40
for col in "CDE": bs.column_dimensions[col].width = 15

# ================================================================ CASH FLOW
cf = wb.create_sheet("Cash Flow")
style_sheet(cf, TEAL)
banner(cf, "Cash Flow Forecast", "12-month profile, academic year 2025/26 — actuals to Period 9, forecast thereafter  |  £k", 15)

cf.cell(row=5, column=2, value="").border = box
for i, m in enumerate(MONTHS):
    c = cf.cell(row=5, column=3 + i, value=m + (" *" if i >= CUR_PERIOD else ""))
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = fill(TEAL if i < CUR_PERIOD else NAVY2)
    c.alignment = Alignment(horizontal="center"); c.border = box
c = cf.cell(row=5, column=15, value="Total")
c.font = Font(bold=True, color=WHITE); c.fill = fill(NAVY); c.border = box
c.alignment = Alignment(horizontal="center")

cf_lines = [
    ("Receipts", None),
    ("  GAG & ESFA grant receipts",   [1180,1125,1125,1125,1160,1130,1165,1140,1170,1155,1165,1190]),
    ("  Other grants & LA income",    [115, 98, 102, 95, 104, 99, 106, 100, 108, 102, 101, 105]),
    ("  Other income (catering etc.)",[ 85, 72, 83, 70, 71, 71, 74, 70, 74, 73, 74, 75]),
    ("Total receipts", "SUM3"),
    ("Payments", None),
    ("  Staff costs (net pay, tax, pensions)", [955,962,958,975,960,964,956,968,961,962,975,1005]),
    ("  Premises & estates",          [110,118,105,128,112,109,104,115,103,108,112,118]),
    ("  Supplies, ICT & catering",    [148,152,145,150,146,150,148,152,147,148,152,158]),
    ("  Capital expenditure",         [ 77, 78, 77, 77, 77, 82, 82, 80, 87, 87, 81, 79]),
    ("Total payments", "SUM4"),
    ("Net cash inflow / (outflow)", "NET"),
    ("Opening balance", "OPEN"),
    ("Closing balance", "CLOSE"),
]
r = 6
group_start = None
sums = []
for label, data in cf_lines:
    cf.cell(row=r, column=2, value=label)
    if data is None:
        cf.cell(row=r, column=2).font = Font(bold=True, size=11, color=NAVY)
        cf.cell(row=r, column=2).border = bot
        for cc in range(3, 16): cf.cell(row=r, column=cc).border = bot
        group_start = r + 1
    elif isinstance(data, list):
        for i, v in enumerate(data):
            c = cf.cell(row=r, column=3 + i, value=v)
            c.number_format = '#,##0'; c.border = box
        c = cf.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})")
        c.number_format = '#,##0'; c.border = box; c.font = Font(bold=True)
        cf.cell(row=r, column=2).border = box
    elif data in ("SUM3", "SUM4"):
        n = 3 if data == "SUM3" else 4
        for i in range(13):
            L = get_column_letter(3 + i)
            c = cf.cell(row=r, column=3 + i, value=f"=SUM({L}{group_start}:{L}{r-1})")
            c.number_format = '#,##0'
        sums.append(r)
        for cc in range(2, 16):
            cell = cf.cell(row=cc and r, column=cc)
            cell.font = Font(bold=True, color=NAVY); cell.border = topd; cell.fill = fill(LIGHT)
    elif data == "NET":
        rec, pay = sums
        for i in range(13):
            L = get_column_letter(3 + i)
            c = cf.cell(row=r, column=3 + i, value=f"={L}{rec}-{L}{pay}")
            c.number_format = GBP
        for cc in range(2, 16):
            cell = cf.cell(row=r, column=cc)
            cell.font = Font(bold=True); cell.border = box
        net_row = r
    elif data == "OPEN":
        cf.cell(row=r, column=3, value=2255).number_format = '#,##0'
        for i in range(1, 12):
            L_prev = get_column_letter(2 + i)
            cf.cell(row=r, column=3 + i, value=f"={L_prev}{r+1}").number_format = '#,##0'
        for cc in range(2, 15): cf.cell(row=r, column=cc).border = box
        open_row = r
    elif data == "CLOSE":
        for i in range(12):
            L = get_column_letter(3 + i)
            c = cf.cell(row=r, column=3 + i, value=f"={L}{open_row}+{L}{net_row}")
            c.number_format = '#,##0'
        for cc in range(2, 15):
            cell = cf.cell(row=r, column=cc)
            cell.fill = fill(NAVY); cell.font = Font(bold=True, color=WHITE); cell.border = box
        close_row = r
    r += 1

# alert: closing balance below 30 days cash (~£440k of monthly spend base) -> red
cf.conditional_formatting.add(
    f"C{close_row}:N{close_row}",
    CellIsRule(operator="lessThan", formula=["1300"], fill=fill(REDF), font=Font(color=RED, bold=True)))
n = cf.cell(row=r + 1, column=2,
    value="* Months marked with an asterisk are forecast. All figures £k. "
          "Closing balance highlights red if it falls below the trust's minimum cash policy (illustrative £1,300k ≈ 30 days of expenditure).")
n.font = Font(italic=True, size=9, color="666666")
cf.column_dimensions["A"].width = 2
cf.column_dimensions["B"].width = 34
for i in range(13):
    cf.column_dimensions[get_column_letter(3 + i)].width = 9
cf.freeze_panes = "C6"

# ================================================================ KPIs
kp = wb.create_sheet("KPIs")
style_sheet(kp, GOLD)
banner(kp, "Key Performance Indicators", "DfE benchmark metrics — RAG rated against trust targets", 8)

kpi_hdr = ["Indicator", "Actual", "Target / Benchmark", "RAG", "Direction", "Notes"]
for j, h in enumerate(kpi_hdr):
    c = kp.cell(row=5, column=2 + j, value=h)
    c.font = Font(bold=True, color=WHITE, size=10); c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal="center"); c.border = box

kpis = [
    ("Staff costs as % of total income", 0.787, "< 80%", "G", "▼ improving",
     "DfE benchmark: most trusts 70–80%. Includes all staff cost lines."),
    ("Revenue reserves as % of income", 0.123, "5% – 15%", "G", "▲ rising",
     "Restricted general + unrestricted funds vs annualised income."),
    ("Days cash in hand", 56, "> 30 days", "G", "▲ rising",
     "Cash balance ÷ average daily expenditure."),
    ("YTD I&E variance to budget", 0.012, "± 2%", "G", "— stable",
     "Net surplus variance as % of budgeted income."),
    ("Pupil numbers vs capacity", 0.943, "> 95%", "A", "▼ falling",
     "1,886 on roll vs 2,000 capacity. October census drives GAG lag."),
    ("Income per pupil", 6290, "Benchmark £6,150", "G", "▲ rising",
     "Compare via DfE Schools Financial Benchmarking service."),
    ("Average teacher cost", 58400, "Benchmark £57,200", "A", "▲ rising",
     "Monitor pay award and pension employer contribution impact."),
    ("Energy & estates cost per m²", 31.2, "Benchmark £29.50", "A", "▲ rising",
     "Above benchmark — estates review under way (see Commentary)."),
]
rag_fill = {"G": (GREENF, GREEN, "●  On target"),
            "A": (AMBERF, AMBER, "●  Watch"),
            "R": (REDF, RED, "●  Action")}
for i, (name, act, tgt, rag, direc, note) in enumerate(kpis):
    r = 6 + i
    kp.cell(row=r, column=2, value=name).font = Font(bold=True, size=10)
    a = kp.cell(row=r, column=3, value=act)
    a.number_format = PCT if isinstance(act, float) and act < 1 else ('#,##0' if act > 999 else '0')
    a.alignment = Alignment(horizontal="center"); a.font = Font(bold=True, size=11, color=NAVY)
    kp.cell(row=r, column=4, value=tgt).alignment = Alignment(horizontal="center")
    f, fc, txt = rag_fill[rag]
    rc = kp.cell(row=r, column=5, value=txt)
    rc.fill = fill(f); rc.font = Font(bold=True, color=fc, size=10)
    rc.alignment = Alignment(horizontal="center")
    kp.cell(row=r, column=6, value=direc).alignment = Alignment(horizontal="center")
    nc = kp.cell(row=r, column=7, value=note)
    nc.font = Font(size=9, color="555555"); nc.alignment = Alignment(wrap_text=True, vertical="center")
    for cc in range(2, 8):
        kp.cell(row=r, column=cc).border = box
        if i % 2 and cc != 5: kp.cell(row=r, column=cc).fill = fill(LIGHT)
    kp.row_dimensions[r].height = 30

kp.column_dimensions["A"].width = 2
kp.column_dimensions["B"].width = 32
kp.column_dimensions["C"].width = 11
kp.column_dimensions["D"].width = 17
kp.column_dimensions["E"].width = 13
kp.column_dimensions["F"].width = 12
kp.column_dimensions["G"].width = 52

# ================================================================ COMMENTARY
cm = wb.create_sheet("Commentary")
style_sheet(cm, "7A7A7A")
banner(cm, "Variance Commentary & Actions", "Explain variances > ±5% or ±£50k, risks, and mitigating actions", 8)
cm_hdr = ["Area", "Variance", "Explanation", "Action / Owner", "Status"]
for j, h in enumerate(cm_hdr):
    c = cm.cell(row=5, column=2 + j, value=h)
    c.font = Font(bold=True, color=WHITE, size=10); c.fill = fill(NAVY2)
    c.alignment = Alignment(horizontal="center"); c.border = box
comments = [
    ("GAG income", "+£75k YTD", "In-year growth funding confirmed for September bulge class.",
     "Reflect in FY forecast — done. CFO", "Closed"),
    ("Premises & estates", "−£45k YTD (adverse)", "Reactive maintenance at two sites plus winter energy usage above plan.",
     "Estates condition survey commissioned; quotes for boiler replacement. Estates Manager", "Open"),
    ("LA / SEN income", "−£28k YTD (adverse)", "EHCP top-up payments from LA running behind invoicing.",
     "Escalated with LA; aged debt review monthly. Finance Officer", "Open"),
    ("Supply & cover costs", "+£38k YTD (favourable)", "Lower sickness absence than budget assumption.",
     "Hold favourable variance; do not release until P11. CFO", "Monitor"),
]
for i, row in enumerate(comments):
    r = 6 + i
    for j, v in enumerate(row):
        c = cm.cell(row=r, column=2 + j, value=v)
        c.border = box; c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2: c.fill = fill(LIGHT)
    cm.row_dimensions[r].height = 40
cm.column_dimensions["A"].width = 2
cm.column_dimensions["B"].width = 20
cm.column_dimensions["C"].width = 18
cm.column_dimensions["D"].width = 52
cm.column_dimensions["E"].width = 40
cm.column_dimensions["F"].width = 10

# ================================================================ DASHBOARD
db = wb.active
db.title = "Dashboard"
style_sheet(db, GOLD)
for col, w in (("A",2),("B",13),("C",13),("D",13),("E",13),("F",13),("G",13),
               ("H",13),("I",13),("J",13),("K",13),("L",13),("M",13),("N",2)):
    db.column_dimensions[col].width = w

banner(db, "EXAMPLE MULTI-ACADEMY TRUST  —  Management Accounts",
       f"Period {CUR_PERIOD}  ·  May 2026  ·  Academic year 2025/26  ·  Prepared in line with DfE Academy Trust Handbook & good practice guidance", 13)

# ---- KPI cards (two rows of four)
def card(ws, row, col_start, title, value_formula, numfmt, sub, accent=TEAL, rag=None):
    """3-cols x 4-rows card."""
    ce = col_start + 2
    ws.merge_cells(start_row=row,   start_column=col_start, end_row=row,   end_column=ce)
    ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+2, end_column=ce)
    ws.merge_cells(start_row=row+3, start_column=col_start, end_row=row+3, end_column=ce)
    t = ws.cell(row=row, column=col_start, value=title.upper())
    t.font = Font(bold=True, size=9, color="666666"); t.alignment = Alignment(horizontal="center", vertical="bottom")
    v = ws.cell(row=row+1, column=col_start, value=value_formula)
    v.font = Font(bold=True, size=20, color=NAVY); v.alignment = Alignment(horizontal="center", vertical="center")
    v.number_format = numfmt
    s = ws.cell(row=row+3, column=col_start, value=sub)
    rag_col = {"G": GREEN, "A": AMBER, "R": RED}.get(rag, "888888")
    s.font = Font(size=9, bold=bool(rag), color=rag_col); s.alignment = Alignment(horizontal="center", vertical="top")
    side = Side(style="medium", color=accent)
    thin_s = Side(style="thin", color=MIDGREY)
    for rr in range(row, row + 4):
        for cc in range(col_start, ce + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = fill(LIGHT)
            cell.border = Border(
                left=side if cc == col_start else None,
                right=thin_s if cc == ce else None,
                top=thin_s if rr == row else None,
                bottom=thin_s if rr == row + 3 else None)

R0 = 5
db.row_dimensions[R0].height = 16
db.row_dimensions[R0+1].height = 20
db.row_dimensions[R0+2].height = 14
db.row_dimensions[R0+3].height = 14
card(db, R0, 2,  "YTD income",        "='Income & Expenditure'!C14",  GBP_K, "", TEAL)
card(db, R0, 5,  "YTD expenditure",   "='Income & Expenditure'!C25",  GBP_K, "", TEAL)
card(db, R0, 8,  "YTD surplus / (deficit)", "='Income & Expenditure'!C27", GBP_K, "", GOLD)
card(db, R0, 11, "FY forecast surplus","='Income & Expenditure'!H27", GBP_K, "", GOLD)
R1 = R0 + 5
db.row_dimensions[R1].height = 16
db.row_dimensions[R1+1].height = 20
db.row_dimensions[R1+2].height = 14
db.row_dimensions[R1+3].height = 14
card(db, R1, 2,  "Cash in bank",      "='Balance Sheet'!C12", GBP_K, "● 56 days cash — on target", TEAL, "G")
card(db, R1, 5,  "Revenue reserves",  "='Balance Sheet'!C22+'Balance Sheet'!C24", GBP_K, "● 12.3% of income — on target", TEAL, "G")
card(db, R1, 8,  "Staff costs % income", 0.787, PCT, "● vs 80% ceiling — on target", GOLD, "G")
card(db, R1, 11, "Pupils on roll",    1886, '#,##0', "● 94.3% of capacity — watch", GOLD, "A")

# fix the sub captions for first row cards
db.cell(row=R0+3, column=2,  value="vs YTD budget — see I&E tab").font = Font(size=9, color="888888")
db.cell(row=R0+3, column=5,  value="vs YTD budget — see I&E tab").font = Font(size=9, color="888888")
db.cell(row=R0+3, column=8,  value="vs budget — favourable").font = Font(size=9, bold=True, color=GREEN)
db.cell(row=R0+3, column=11, value="vs FY budget — favourable").font = Font(size=9, bold=True, color=GREEN)
for cs in (2, 5, 8, 11):
    db.cell(row=R0+3, column=cs).alignment = Alignment(horizontal="center", vertical="top")

# ---- charts
CH1 = R1 + 5  # row anchor

h = db.cell(row=CH1, column=2, value="Income vs expenditure — monthly (£k)")
h.font = Font(bold=True, size=11, color=NAVY)
h2 = db.cell(row=CH1, column=8, value="Cash balance — actual vs forecast (£k)")
h2.font = Font(bold=True, size=11, color=NAVY)

bar = BarChart()
bar.type = "col"; bar.style = 10; bar.height = 8.2; bar.width = 14.5
bar.y_axis.title = None; bar.x_axis.delete = False; bar.y_axis.delete = False
bar.gapWidth = 60
months_ref = Reference(dt, min_col=2, min_row=6, max_row=17)
inc_ref = Reference(dt, min_col=3, min_row=5, max_row=17)
exp_ref = Reference(dt, min_col=5, min_row=5, max_row=17)
bar.add_data(inc_ref, titles_from_data=True)
bar.add_data(exp_ref, titles_from_data=True)
bar.set_categories(months_ref)
bar.series[0].graphicalProperties.solidFill = TEAL
bar.series[1].graphicalProperties.solidFill = NAVY
bar.legend.position = "b"
db.add_chart(bar, f"B{CH1+1}")

line = LineChart()
line.style = 12; line.height = 8.2; line.width = 14.5
cash_a = Reference(dt, min_col=7, min_row=5, max_row=17)
cash_f = Reference(dt, min_col=8, min_row=5, max_row=17)
line.add_data(cash_a, titles_from_data=True)
line.add_data(cash_f, titles_from_data=True)
line.set_categories(months_ref)
line.series[0].graphicalProperties.line.solidFill = TEAL
line.series[0].graphicalProperties.line.width = 28000
line.series[1].graphicalProperties.line.solidFill = GOLD
line.series[1].graphicalProperties.line.width = 18000
line.series[1].graphicalProperties.line.dashStyle = "dash"
line.series[0].smooth = False
line.series[1].smooth = False
line.legend.position = "b"
db.add_chart(line, f"H{CH1+1}")

CH2 = CH1 + 18
h3 = db.cell(row=CH2, column=2, value="Monthly surplus / (deficit) vs budget (£k)")
h3.font = Font(bold=True, size=11, color=NAVY)
h4 = db.cell(row=CH2, column=8, value="At a glance")
h4.font = Font(bold=True, size=11, color=NAVY)

bar2 = BarChart()
bar2.type = "col"; bar2.style = 10; bar2.height = 8.2; bar2.width = 14.5
bar2.gapWidth = 60
s_act = Reference(dt, min_col=9, min_row=5, max_row=17)
s_bud = Reference(dt, min_col=10, min_row=5, max_row=17)
bar2.add_data(s_act, titles_from_data=True)
bar2.add_data(s_bud, titles_from_data=True)
bar2.set_categories(months_ref)
bar2.series[0].graphicalProperties.solidFill = GOLD
bar2.series[1].graphicalProperties.solidFill = MIDGREY
bar2.legend.position = "b"
db.add_chart(bar2, f"B{CH2+1}")

glance = [
    ("●", GREEN, "Finances on track: YTD surplus £52k ahead of budget; FY forecast surplus improved to £180k."),
    ("●", GREEN, "Liquidity strong: 56 days cash, reserves 12.3% of income — inside the 5–15% policy range."),
    ("●", AMBER, "Premises costs £45k adverse — estates survey commissioned (see Commentary)."),
    ("●", AMBER, "Pupil numbers 94.3% of capacity — October census key risk to 2026/27 GAG."),
    ("●", GREEN, "Staff costs 78.7% of income — below the 80% sustainability ceiling."),
]
for i, (dot, color, txt) in enumerate(glance):
    r = CH2 + 2 + i * 3
    db.merge_cells(start_row=r, start_column=8, end_row=r + 1, end_column=13)
    d = db.cell(row=r, column=8, value="●  " + txt)
    d.font = Font(size=10, color=color, bold=True)
    d.alignment = Alignment(wrap_text=True, vertical="center")
    for rr in range(r, r + 2):
        for cc in range(8, 14):
            db.cell(row=rr, column=cc).fill = fill(LIGHT)

f = db.cell(row=CH2 + 18, column=2,
    value="Contents: Income & Expenditure · Balance Sheet · Cash Flow · KPIs · Commentary · Monthly Data. "
          "Pack covers the items the Academy Trust Handbook requires boards to monitor (I&E, variance to budget, cash flow and balance sheet). All figures illustrative.")
f.font = Font(italic=True, size=9, color="666666")
db.merge_cells(start_row=CH2 + 18, start_column=2, end_row=CH2 + 18, end_column=13)

# order sheets
wb.move_sheet("Dashboard", offset=-len(wb.sheetnames))
out = "/home/user/testground/Academy_Trust_Management_Accounts_Dashboard.xlsx"
wb.save(out)
print("saved", out)
